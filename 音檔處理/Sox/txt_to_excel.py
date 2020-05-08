#讀txt檔處理(換行問題)
#逐值載入
#若 key(頻率) 相同 則 value(振幅) 相加
#累加sum,用來平均用
#平均振幅
#匯出excel檔案
#used.txt用來判斷是否已經轉檔

import os
import xlwt
import openpyxl
from openpyxl import load_workbook,Workbook


#處理txt檔(用dict存值，平均)
def txt_to_dict(f,dic,filename):
    #累加(用來平均用)
    sum = 0
    #判斷頻率或是振幅
    num = 1
    #逐值載入
    for i in f.split('  '):
        #終止條件
        if i == '':
            break
        #分頻率(1)跟振幅(0)
        #頻率
        if num % 2 == 1 :
            #去小數點
            a = int(float(i))
            if a == 0:
                sum += 1
            #是否存在此頻率
            if a in dic : 
                num += 1
            else:
                dic[a] = 0
                num += 1
        #振幅
        else:
            c = float(i)
            dic[a] = dic[a]+c
            num += 1
    
    return dic

#各檔案 按照頻率列出全部的振幅
def savesameFreq(f,dic,filename):
    savedic = {}
    #判斷頻率或是振幅
    num = 1
    for i in f.split('  '):
        
        #終止條件
        if i == '':
            break
        
        #頻率
        if num % 2 == 1:
            tmpfreq = i
            #print('頻率')
            if int(float(tmpfreq)) in savedic:
                num += 1
            else:
                savedic[int(float(tmpfreq))] = []
                #print(savedic)
                num += 1
        #振幅
        else:
            #print('振幅')
            savedic[int(float(tmpfreq))].append(i)
            num += 1
            
    #儲存進excel
    xls=load_workbook("C:/Users/3c/Desktop/music/savesameFreq.xlsx")    
    sheet = xls.get_sheet_by_name('Sheet') #生成excel的方法
    #xls = load_workbook()  
    #ws = xls.active
    row ,column = 1,1
    tmp = 'A'+str(row)
    #判斷全新一列 塞值
    while True:
        #此儲存格是否為空來判斷是否塞值
        if sheet[tmp].value != None:
            row += 1
            tmp = 'A'+str(row)
            #print(ws[tmp].value)
        else:
            break
    #塞入檔案名稱
    sheet.cell(row,column,value=filename[0:-4])
    row += 1
    #依序取頻率
    for ansFreq in savedic:
        if ansFreq > 700:
            break
        #加總
        sum = 0
        #平均用
        count = 1
        column = 1
        sheet.cell(row,column,value="頻率:"+str(ansFreq))
        column = 4
        #取出此頻率的全部振幅
        for ansAmplitude in savedic[ansFreq]:
            sum += float(ansAmplitude)
            sheet.cell(row,column,value=ansAmplitude)
            column += 1
            count += 1
        #塞入加總值
        sheet.cell(row,2,value='加總:'+str(sum))
        column += 1
        #塞入平均值
        sheet.cell(row,3,value='平均:'+str(sum/count))
        row += 1
    #存excel
    xls.save("C:/Users/3c/Desktop/music/savesameFreq.xlsx") 
    print(str(filename[0:-4])+' \n振幅前十大資料已儲存進:savasameFreq_test.xlsx\n')
    
#匯出excel 頻率對應振幅前十大
def dict_to_excel_MaxTen(f,dic,filename):
    tmpdic = dic
    maxlis = {}
    #判斷方法: 找出tmpdic 內value最大的值
    #          將振幅放置maxlis中 maxlis[tmp] = tmplis[tmp]
    #          將tmplis中的tmp pop掉 =>最大值改變
    
    for i in range(0,10):
        tmp  = max(tmpdic,key=tmpdic.get)
        maxlis[tmp] = tmpdic[tmp]                                                                   
        tmpdic.pop(tmp)
    #print(maxlis)
    
    xls = load_workbook("C:/Users/3c/Desktop/music/MaxTen_EachSegment.xlsx")
    ws = xls.active
    row,column = 1,1
    tmp = 'A'+str(row)
    #判斷全新一列 塞值
    while True:
        if ws[tmp].value != None:
            row += 1
            tmp = 'A'+str(row)
            #print(ws[tmp].value)
        else:
            break
    #檔案名稱
    ws.cell(row,column,value=filename[0:-4])
    row += 1
    for i in maxlis:
        ws.cell(row,column,value='頻率:'+str(i))
        column += 1
        ws.cell(row,column,value='振幅:'+str(maxlis[i]))
        row += 1
        column = 1
    xls.save("C:/Users/3c/Desktop/music/MaxTen_EachSegment.xlsx") 
    print(str(filename[0:-4])+' \n各頻率資料已儲存進:MaxTen_EachSegment.xlsx\n')
    
#匯出成excel檔案(all)
def dict_to_excel(f,dic,filename,store_position,store_xlsx):

    # 頻率取0~700
    tmpdic = {}
    for j in dic :
        if j > 700:
            break
        else:
            tmpdic[j] = dic[j]
    # 頻率取190~550
    """
    for j in dic:
        if j < 190:
            continue
        if j > 550:
            break
        tmpdic[j] = dic[j]
    """
    
    tmpfreq = max(tmpdic,key=tmpdic.get)
    # maxinterval:取振福最大值用來正規畫
    maxinterval = tmpdic[tmpfreq]

    row,column = 1,3
    
    #頻率間隔數值(EX:interval=5 or interval=10)
    for tmp in dic:
        if tmp != 0:
            freqinterval = tmp
            break
    
    savename = store_xlsx

    xls=load_workbook(savename)
    ws = xls.active
    tmp = 'A'+str(row)
    #判斷新一列 將值塞入
    while True:
        if ws[tmp].value != None:
            row += 1
            tmp = 'A'+str(row)
            #print(ws[tmp].value)
        else:
            break
    
    #塞音檔名稱
    ws.cell(row,1,value=filename[0:-4])
    #塞蜂巢名稱
    ws.cell(row,2,value=filename[0:5])
    #塞頻率值
    freqrow,freqcolumn = 1,3
    #頻率控制在700以下
    for freq in dic:
        #頻率取0~700
        if freq > 700:
            break
        ws.cell(freqrow,freqcolumn,value=freq)
        freqcolumn += 1
        # 頻率取190~550
        """
        if freq >=190 and freq <= 550:
            ws.cell(freqrow,freqcolumn,value=freq)
            freqcolumn += 1
        if freq > 550:
            break
        """
    
    for i in dic:
        #頻率:700以前塞值
        if i >700:
            break
        #print(i,dic[i])
        #加入振幅(已經做過正規畫)
        ws.cell(row,column,value=dic[i]/maxinterval)
        #excel換行
        column += 1
        """
        if i >= 190 and i <= 550:
        #print(i,dic[i])
        #加入振幅   
            #print(dic[i])
            ws.cell(row,column,value=dic[i])
        #excel換行
            column += 1
        if i > 550:
            break
        """
        
    row += 1
    xls.save(savename)

# 確定此刻音檔資料是否已經使用過
def check(store_xlsx):
    checklis = []
    if not os.path.isfile(store_xlsx): # xlsx尚未建立過 -> 不需檢查 
        xlsx = openpyxl.Workbook()
        sheet = xlsx.get_sheet_by_name('Sheet')
        sheet.cell(row=1,column=1,value='Name')
        sheet.cell(row=1,column=2,value='Hive')
        xlsx.save(store_xlsx)
    else:
        xlsx=load_workbook(store_xlsx)
        sheet = xlsx.active
        for row in sheet.rows:
            checklis.append(row[0].value)
    return checklis

    
def main():
    
    file_position = os.getcwd() + os.sep +  'sound_data' + os.sep # 要處理的音檔資料位置
    file_List = os.listdir(file_position)  # 取得音檔資料
    
    store_position = os.getcwd() + os.sep
    store_xlsx = store_position + 'sox.xlsx'
    
    # 從之前的csv取得data 確定此刻資料是否使用過
    checklis = check(store_xlsx) # 還沒放進去判斷 
    
    # 依序取出作處理
    for filename in file_List:
        # 判斷是否使用過
        if filename[0:-4] in checklis:
            print(filename + ' has been used')
            continue
        print('檔案名稱:'+str(filename))
        if filename.endswith(".txt"):
            openfile = os.getcwd() + os.sep + 'sound_data' + os.sep + filename
            w = open(openfile).read()
            f = w.replace('\n','  ')
            dic = {}
            #單一檔案 按照頻率列出全部的振幅
                
            txt_to_dict(f,dic,filename)
            #全部值
            dict_to_excel(f,dic,filename,store_position,store_xlsx)
            #savesameFreq(f,dic,filename)
                
            #segmentstat(f,dic,filename)

if __name__ == "__main__":
    main()

"""
終止條件問題解決:
取值方法 -> 以 ('  ') 為取出規則
最後會取到空值 -> 產生無法str轉換成int的問題
"""