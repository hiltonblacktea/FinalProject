import os
import librosa
import librosa.display
import csv
import json
import re
import numpy as np
import random
import IPython.display as ipd
from sklearn import preprocessing,svm
from info import i, printb, printr, printp, print
from numpy import *
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import sys
import glob
def load_audioFiles_saves_segments(audiofilenames_list, path_audioFiles,path_save_audio_labels, block_size, save_audioSegments='yes'):
    printb("Number of audiofiles in folder: "+str(len(audiofilenames_list)))
    fi=0 #目前檔案第0個
    #操作方式: 逐檔載入,一個音檔中設置開始時間offset從0開始,到指定每個音檔分段時間duration(block_size=60s),然後進到while 1,
    #將音檔輸出成wav(等於此音檔的序列被輸出,剩下的序列是從offset+block_size開始)&資訊存進csv後,序列還有資料就再執行一次,
    #當序列<0則表示此音檔已經完全輸出,則跳出while 1 ,繼續處理下一個音檔
    
    for file_name in audiofilenames_list:
        fi=fi+1 
        print('\n')
        printb('Processing '+ file_name+'       :::file number:  '+str(fi)+' ----->of '+str(len(audiofilenames_list)))
          
        offset=0
        block_id =0
        # 處理前先判斷是否已經處理過,若以處理過則跳過
        exitfile = fileisexitornot(path_save_audio_labels)
        # 用第一個切割的音段名稱來判斷是否需要處理此音檔
        first_block_name = file_name[0:-4] + '__segment'+str(block_id) + '.wav'
        if first_block_name in exitfile:
            print('this file has been used')
            continue

        while 1:
            
            # READ ONE BLOCK OF THE AUDIO FILE
            try: #try除錯用
                #加載音檔 (音檔位置,offset:以此時間開始讀取,duration:持續讀取時間) 回傳音檔序列,sr=採樣率
                block,sr = librosa.core.load(path_audioFiles+file_name, offset=offset, duration=block_size)
                print('-----Reading segment '+str(block_id))
            except ValueError as e:
                e
                if 'Input signal length' in str(e):
                    block=np.arange(0)
            except FileNotFoundError as e1:
                print(e1, ' but continuing anyway')
                
            if block.shape[0] > 0:    
            #when total length = multiple of blocksize, results that last block is 0-lenght, this if bypasses those cases.
                block_name=file_name[0:-4]+'__segment'+str(block_id)
                print(block_name)
                
                
                #testing
                #在此產生state_label.csv並將名稱跟label狀態加入
                states=['active','missing queen','swarm' ]
                
                label_file_exists = os.path.isfile(path_save_audio_labels+'state_labels'+'.csv') #判斷檔案是否存在
                with open(path_save_audio_labels+'state_labels'+'.csv','a',newline='') as label_file: #label_file檔案讀寫
                    writer = csv.DictWriter(label_file,fieldnames=['sample_name','label'],delimiter=',') #定義欄位sample_name, label
                    if not label_file_exists: #假如檔案不存在
                        writer.writeheader()  #寫入欄位
                    
                    csvreader = csv.DictReader(label_file)
                    try:
                        for row in csvreader:
                            print(row['sample_name'])
                    except:
                        pass
                    #270
                    label_state=read_HiveState_fromSampleName(block_name, states)
                    writer.writerow({'sample_name':block_name, 'label':label_state})
                # READ BEE NOT_BEE ANNOTATIONS: 將音檔資訊.lab儲存進csv
                
                # MAKE BLOCK OF THE SAME SIZE:
                if block.shape[0] < block_size*sr:   
                    pass

                        
            
                # Save audio segment:
                if save_audioSegments=='yes' and (not os.path.exists(path_save_audio_labels+block_name+'.wav')): #saves only if option is chosen and if block file doesn't already exist.
                    #將音檔序列輸出為.wav文件
                    librosa.output.write_wav(path_save_audio_labels+block_name+'.wav', block, sr)
                    print( '-----Saved wav file for segment '+str(block_id))
                
                    
                    
            else :
                print('-----no more segments for this file-----')
                print('\n')
                break
            offset += block_size
            block_id += 1
    printb('_____No more audioFiles_____')
       
    return 
    #從名稱得知音檔label 回傳值為label狀態
    
# 將切割完的音檔名稱存到.json檔中
def write_sample_ids_perHive(sample_ids , savepath):
    uniqueHivesNames={}
    for sample in sample_ids:
        if 'Hive' in uniqueHivesNames.keys():
            uniqueHivesNames['Hive'].append(sample)
        else:
            uniqueHivesNames['Hive'] = [sample]
    with open(savepath+'sampleID_perHive.json', 'w') as outfile:
        json.dump(uniqueHivesNames, outfile)
        
    return uniqueHivesNames
   

# 判定蜂巢狀態
def read_HiveState_fromSampleName(filename, states):   #states: state_labels=['active','missing queen','swarm' ]
    label_state='other'
    for state in states:
        if state in filename.lower():
            label_state = state
    #incorporate condition for Nu-hive recordings which do not follow the same annotation: 'QueenBee' or 'NO_QueenBee'        
    if label_state=='other':
        if 'NO_QueenBee' in filename:
            label_state = states[1]
        else:
            label_state=states[0]
    return label_state
    
    #將label存入state_label.csv,方法是找尋檔案名稱內所對應的label(ex:GH001 - Active - Day - 141022_0659_0751__segment5 ==> 找到label=Active)
    
    
        
    
# 從.json檔案取出data
def get_samples_id_perSet(pathSplitFile):  # reads split_id file
    split_dict=json.load(open (pathSplitFile, 'r'))
    #轉換的音檔按照此處規定的蜂巢來處理
    sampleID = split_dict['Hive']
    return sampleID

# MFCC特徵提取並存至excel中
def get_features_from_samples(path_audio_samples, sample_ids, raw_feature, normalization, high_level_features ): #normalization = NO, z_norm, min_max
    ## function to extract features 
    ## 提取特徵
    
    dic = {}
    
    for sample in sample_ids:
        
        # 原始特徵提取(MFCC序列)
        x = raw_feature_fromSample(path_audio_samples+sample, raw_feature ) # x.shape: (4, 20, 2584)
        
        # 正規化
        if not normalization == 'NO': #是否要正規化
            
            x_norm = featureMap_normalization_block_level(x, normalizationType = normalization) 
                     
        else: 
            
            x_norm = x
        
        if high_level_features:
            # high level feature extraction:
            if 'MFCCs' in raw_feature:
                X = compute_statistics_overMFCCs(x_norm, 'yes') # X.shape: (4 , 120)
                    # 864
            else: 
                X = compute_statistics_overSpectogram(x_norm)
                    # 853
            feature_map=X
        else:
            feature_map=x_norm
            
        #取前五個字用來區分不同蜂巢(ex:檔案名稱=Hive3_20_07_2017_QueenBee_H3_audio___06_20_00__segment6.wav)
        tmpdataname = sample[0:5]
        
        dic[sample] = list(feature_map)
    
    StoreExcel(dic)
    """
    # 取出之前data 避免重複處理
    checklis = []
    
    if  os.path.isfile('./MFCC/Name_Label_MFCCfeature.xlsx'):
        print("file existed ")
        try:
            df = pd.read_excel('./MFCC/Name_Label_MFCCfeature.xlsx',delimiter="\t")
            # 取出第一行(Name)
            check = df.iloc[:,0]
            for i in check:
                if i in checklis:
                    continue
                else:
                    checklis.append(i)
        except Exception as e:
            print("An error found but ignored",e)
    #以下是儲存進excel
    xls=openpyxl.Workbook()
    sheet = xls.get_sheet_by_name('Sheet') #生成excel的方法
    x,y = 1,1
    
    sheet.cell(row=x,column=y,value='Name')
    sheet.cell(row=x,column=y+1,value='Labal')
    
    
    x ,y = 2 , 1
    
    #判斷全新一列 塞值
    tmp = 'A'+str(x)
    while True:
        if sheet[tmp].value != None:
            x += 1
            tmp = 'A'+str(x)
            #print(ws[tmp].value)
        else:
            break
    
    for name in dic :
        if name[0:-4] in checklis :
            print('this file has been used , Pass!')
            continue # 此檔案已經存過 略過
        
        currentname = name[0:-4]
        
        # 判斷狀況
        if 'NO' in currentname  or 'Missing' in currentname:
            currentlabel = 'Missing Queen'
        else:
            currentlabel = 'Active'
        
        # 存進excel格子
        # 存 名稱 & 狀態
        y = 1
        sheet.cell(row=x,column=y,value=currentname)
        sheet.cell(row=x,column=y+1,value=currentlabel)
        y = 3
        # 存 feature
        
        for feature in dic[name]:
            sheet.cell(row=x,column=y,value=feature)
            y += 1
        x += 1

    ## MFCC特徵數值
    x , y =  1,3
    for count in range(1,121):
        tmps = count
        sheet.cell(row=x,column=y,value=tmps)
        y += 1
    
    Excelname = 'Name_Label_MFCCfeature.xlsx'
    xls.save('./MFCC/' + Excelname) 
    print(Excelname+" have Done")
    """
    
    
# MFCC原始序列提取
def raw_feature_fromSample( path_audio_sample, feature2extract ):

    #取得音頻時間序列,採樣率(預設22050) 
    #https://www.google.com.tw/url?sa=t&rct=j&q=&esrc=s&source=web&cd=1&cad=rja&uact=8&ved=2ahUKEwjA4pusptPnAhUqxosBHSn7BpgQFjAAegQIARAB&url=https%3A%2F%2Flibrosa.github.io%2Flibrosa%2Fgenerated%2Flibrosa.core.load.html&usg=AOvVaw3DdAgYRN_GVa4-a30QW6lR
    
    audio_sample, sr = librosa.core.load(path_audio_sample)
    
    #re : 正則表達式 , re.match => 匹配字符串 回傳match object
    m = re.match(r"\w+s(\d+)", feature2extract) # feature2extract = 'MFCCs20'
    n_freqs=int(m.groups()[0])
    Melspec = librosa.feature.melspectrogram(audio_sample, n_mels = n_freqs) # computes mel spectrograms from audio sample, 
                                                                             # 從音頻樣本計算mel譜圖
    if 'LOG' in feature2extract: #'LOG_MELfrequencies48'
        Melspec=librosa.feature.melspectrogram(audio_sample, sr=sr, n_mels=n_freqs)
        x=librosa.power_to_db(Melspec+1)
    elif 'MFCCs' in feature2extract:
        n_freqs = int(feature2extract[5:len(feature2extract)])
        Melspec = librosa.feature.melspectrogram(audio_sample, sr=sr)
        
        # 聲音檔時間序列
        x = librosa.feature.mfcc(S=librosa.power_to_db(Melspec),sr=sr, n_mfcc = n_freqs)
    else:
        x = Melspec
    
    return x   


def compute_statistics_overMFCCs(MFCC, first='yes'):
    x_delta=librosa.feature.delta(MFCC)
    x_delta2=librosa.feature.delta(MFCC, order=2)
    
    if first=='no':
        MFCC=MFCC[1:]
        x_delta=x_delta[1:]
        x_delta2=x_delta2[1:]
    
    X_4features=np.concatenate((np.mean(MFCC,1), np.std(MFCC,1),np.mean(x_delta,1), np.std(x_delta,1), np.mean(x_delta2,1), np.std(x_delta2,1)), axis=0)
    X_flat = np.asarray(X_4features)
    
    return X_flat


def featureMap_normalization_block_level(feature_map, normalizationType='min_max'):
      
    # TODO other levels of normalization (example: whole dataset, set (train, val or test) level)
    if normalizationType== 'min_max': # min_max scaling
        
        min_max_scaler = preprocessing.MinMaxScaler()
        normalized_featureMap = min_max_scaler.fit_transform(feature_map)
    
    if normalizationType == 'z_norm': # standardization(z-normalization)
        normalized_featureMap = preprocessing.scale(feature_map)   

    return normalized_featureMap

def getmusicposition():
    a = os.getcwd().split(os.sep)[-1]
    now = os.getcwd().replace(a,'')
    return now

def fileisexitornot(path_save_audio_labels):
    file = [os.path.basename(x) for x in glob.glob(path_save_audio_labels+'*'+'.wav')] 
    return file

def StoreExcel(dic):
    checklis = []
    
    if  os.path.isfile('./MFCC/Name_Label_MFCCfeature.xlsx'):
        xls=load_workbook('./MFCC/Name_Label_MFCCfeature.xlsx')
        print("file existed ")
        try:
            df = pd.read_excel('./MFCC/Name_Label_MFCCfeature.xlsx',delimiter="\t")
            # 取出第一行(Name)
            check = df.iloc[:,0]
            for i in check:
                if i in checklis:
                    continue
                else:
                    checklis.append(i)
        except Exception as e:
            print("An error found but ignored",e)
    else:
        xls=openpyxl.Workbook()
    #以下是儲存進excel
    
    sheet = xls.get_sheet_by_name('Sheet') #生成excel的方法
    x,y = 1,1
    
    sheet.cell(row=x,column=y,value='Name')
    sheet.cell(row=x,column=y+1,value='Labal')
    
    
    x ,y = 2 , 1
    
    #判斷全新一列 塞值
    tmp = 'A'+str(x)
    while True:
        if sheet[tmp].value != None:
            x += 1
            tmp = 'A'+str(x)
            #print(ws[tmp].value)
        else:
            break
    
    for name in dic :
        if name[0:-4] in checklis :
            print('this file has been used , Pass!')
            continue # 此檔案已經存過 略過
        
        currentname = name[0:-4]
        
        # 判斷狀況
        if 'NO' in currentname  or 'Missing' in currentname:
            currentlabel = 'Missing Queen'
        else:
            currentlabel = 'Active'
        
        # 存進excel格子
        # 存 名稱 & 狀態
        y = 1
        sheet.cell(row=x,column=y,value=currentname)
        sheet.cell(row=x,column=y+1,value=currentlabel)
        y = 3
        # 存 feature
        
        for feature in dic[name]:
            sheet.cell(row=x,column=y,value=feature)
            y += 1
        x += 1

    # MFCC特徵數值
    x , y =  1,3
    for count in range(1,121):
        tmps = count
        sheet.cell(row=x,column=y,value=tmps)
        y += 1
    
    Excelname = 'Name_Label_MFCCfeature.xlsx'
    xls.save('./MFCC/' + Excelname) 
    print(Excelname+" have Done")